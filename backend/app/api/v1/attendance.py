import fastapi
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional
from datetime import date
import uuid
from app.api.v1.dependencies import get_db, get_current_user
from app.schemas.attendance import (
    AttendancePunch, AttendanceRecordRead, AttendanceRecordCreate, AttendanceRecordUpdate,
    DailyAttendanceRead, 
    ShiftRead, ShiftCreate, ShiftUpdate,
    ShiftRosterCreate, ShiftRosterRead, RosterBulkUpdate,
    ShiftBreakCreate, ShiftBreakRead,
    RosterAutoGenerate, RosterCellUpdate, RosterReadEnriched,
    PublicHolidayCreate, PublicHolidayRead, PublicHolidayUpdate, PublicHolidaySeed,
    RosterClear
)
from app.services.attendance import AttendanceService
from app.models.auth import User

router = APIRouter(prefix="/attendance", tags=["Attendance"])

@router.post("/punch", response_model=AttendanceRecordRead)
async def punch(
    punch_data: AttendancePunch,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Verify entity access
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(punch_data.entity_id, current_user, db)
    
    service = AttendanceService(db)
    try:
        record = await service.punch(
            entity_id=punch_data.entity_id,
            employment_id=punch_data.employment_id,
            punch_type=punch_data.punch_type,
            timestamp=punch_data.timestamp,
            source=punch_data.source
        )
        await db.commit()
        return record
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/process-daily")
async def process_daily(
    entity_id: uuid.UUID,
    work_date: date,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    # Only Admin or HR Manager can trigger processing
    if not current_user.is_tenant_admin:
        from app.api.v1.dependencies import get_entity_access
        role = await get_entity_access(entity_id, current_user, db)
        if role != "hr_admin":
            raise HTTPException(status_code=403, detail="Not enough permissions")
            
    service = AttendanceService(db)
    await service.compute_daily_attendance(entity_id, work_date)
    await db.commit()
    return {"message": f"Successfully processed attendance for {work_date}"}

@router.get("/daily-attendance", response_model=List[DailyAttendanceRead])
async def get_daily_attendance(
    entity_id: uuid.UUID,
    start_date: date,
    end_date: date,
    employment_id: Optional[uuid.UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.attendance import DailyAttendance, Shift
    from app.models.employment import Employment, Person
    from sqlalchemy import and_, select
    
    query = (
        select(DailyAttendance, Person.full_name, Shift.name)
        .join(Employment, DailyAttendance.employment_id == Employment.id)
        .join(Person, Employment.person_id == Person.id)
        .outerjoin(Shift, DailyAttendance.scheduled_shift_id == Shift.id)
        .where(
            and_(
                DailyAttendance.entity_id == entity_id,
                DailyAttendance.work_date >= start_date,
                DailyAttendance.work_date <= end_date
            )
        )
    )
    if employment_id:
        query = query.where(DailyAttendance.employment_id == employment_id)
        
    result = await db.execute(query)
    rows = result.all()
    
    summaries = []
    for daily, emp_name, shift_name in rows:
        summary = DailyAttendanceRead.model_validate(daily)
        summary.employee_name = emp_name
        summary.shift_name = shift_name
        summaries.append(summary)
        
    return summaries

@router.get("/current/{employment_id}", response_model=Optional[AttendanceRecordRead])
async def get_current_attendance(
    employment_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.attendance import AttendanceRecord
    from sqlalchemy import and_, select
    from datetime import date
    
    today = date.today()
    query = select(AttendanceRecord).where(
        and_(
            AttendanceRecord.employment_id == employment_id,
            AttendanceRecord.work_date == today
        )
    )
    result = await db.execute(query)
    return result.scalar_one_or_none()

# --- Raw Records Management (Admin) ---

@router.get("/records", response_model=List[AttendanceRecordRead])
async def list_attendance_records(
    entity_id: uuid.UUID,
    start_date: date,
    end_date: date,
    employment_id: Optional[uuid.UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(entity_id, current_user, db)
    
    service = AttendanceService(db)
    return await service.get_attendance_records(entity_id, start_date, end_date, employment_id)

@router.get("/import/template")
async def download_timesheet_template(
    current_user: User = Depends(get_current_user)
):
    from fastapi.responses import Response
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet Import"

    headers = [
        "Employee ID",
        "Date",          # YYYY-MM-DD
        "Clock In",      # HH:MM or YYYY-MM-DD HH:MM
        "Clock Out",     # HH:MM or YYYY-MM-DD HH:MM
        "Notes"
    ]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header) + 5, 15)

    # Sample rows
    samples = [
        ["EMP-001", "2023-10-25", "09:00", "18:00", "Regular shift"],
        ["EMP-002", "2023-10-25", "2023-10-25 08:45", "2023-10-25 17:30", "Early clock in"]
    ]
    for r_idx, row in enumerate(samples, 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Instructions sheet
    note_ws = wb.create_sheet("Instructions")
    instructions = [
        "How to use this template:",
        "1. Employee ID is required and must match the system's Employee ID exactly.",
        "2. Date must be in YYYY-MM-DD format (e.g., 2023-10-25).",
        "3. Clock In/Out can be just time (HH:MM) or full datetime (YYYY-MM-DD HH:MM).",
        "4. If only time is provided, the Date column will be used for the date part.",
        "5. Do not modify the header row (Row 1)."
    ]
    for i, note in enumerate(instructions, 1):
        note_ws.cell(row=i, column=1, value=note)

    # Save to memory
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return Response(
        content=stream.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=timesheet_template.xlsx"
        }
    )

from app.schemas.attendance import TimesheetPreviewResponse, TimesheetConfirmPayload, TimesheetConfirmResponse

@router.post("/import/preview", response_model=TimesheetPreviewResponse)
async def preview_timesheet(
    entity_id: uuid.UUID = Query(...),
    file: fastapi.UploadFile = fastapi.File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    import fastapi
    from app.api.v1.dependencies import get_entity_access
    
    # Must be HR Admin or full admin
    if not current_user.is_tenant_admin:
        role = await get_entity_access(entity_id, current_user, db)
        if role != "hr_admin":
            raise HTTPException(status_code=403, detail="Not enough permissions to import timesheets")
            
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV or Excel files are accepted")
        
    content = await file.read()
    service = AttendanceService(db)
    
    try:
        result = await service.preview_timesheet_data(entity_id, content, file.filename)
        # Note: Do NOT commit here, this is just a preview
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to process file: {str(e)}")

@router.post("/import/confirm", response_model=TimesheetConfirmResponse)
async def confirm_timesheet(
    payload: TimesheetConfirmPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    
    # Must be HR Admin or full admin
    if not current_user.is_tenant_admin:
        role = await get_entity_access(payload.entity_id, current_user, db)
        if role != "hr_admin":
            raise HTTPException(status_code=403, detail="Not enough permissions to import timesheets")

    service = AttendanceService(db)
    try:
        # Convert Pydantic models to dicts for the service method
        records_data = [rec.model_dump() for rec in payload.records]
        result = await service.confirm_timesheet_import(payload.entity_id, records_data)
        await db.commit()
        return result
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to confirm import: {str(e)}")


@router.post("/records", response_model=AttendanceRecordRead)
async def create_manual_record(
    data: AttendanceRecordCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(data.entity_id, current_user, db)
    
    service = AttendanceService(db)
    record = await service.create_manual_punch(data)
    await db.commit()
    return record

@router.patch("/records/{record_id}", response_model=AttendanceRecordRead)
async def update_attendance_record(
    record_id: uuid.UUID,
    data: AttendanceRecordUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    # Check access (need to fetch record first to get entity_id)
    from app.models.attendance import AttendanceRecord
    from sqlalchemy import select
    stmt = select(AttendanceRecord).where(AttendanceRecord.id == record_id)
    result = await db.execute(stmt)
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(record.entity_id, current_user, db)
    
    updated = await service.update_attendance_record(record_id, data)
    await db.commit()
    return updated

@router.delete("/records/{record_id}")
async def delete_attendance_record(
    record_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    from app.models.attendance import AttendanceRecord
    from sqlalchemy import select
    stmt = select(AttendanceRecord).where(AttendanceRecord.id == record_id)
    result = await db.execute(stmt)
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
        
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(record.entity_id, current_user, db)
    
    await service.delete_attendance_record(record_id)
    await db.commit()
    return {"message": "Record deleted successfully"}

# --- Shifts ---

@router.post("/shifts", response_model=ShiftRead)
async def create_shift(
    shift_in: ShiftCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(shift_in.entity_id, current_user, db)
    
    service = AttendanceService(db)
    shift = await service.create_shift(shift_in)
    await db.commit()
    return shift

@router.get("/shifts", response_model=List[ShiftRead])
async def list_shifts(
    entity_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(entity_id, current_user, db)
    
    service = AttendanceService(db)
    return await service.get_shifts(entity_id)

@router.patch("/shifts/{shift_id}", response_model=ShiftRead)
async def update_shift(
    shift_id: uuid.UUID,
    shift_update: ShiftUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    shift = await service.get_shift(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
        
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(shift.entity_id, current_user, db)
    
    updated_shift = await service.update_shift(shift_id, shift_update)
    await db.commit()
    return updated_shift

@router.delete("/shifts/{shift_id}")
async def delete_shift(
    shift_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    shift = await service.get_shift(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
        
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(shift.entity_id, current_user, db)
    
    await service.delete_shift(shift_id)
    await db.commit()
    return {"message": "Shift deleted successfully"}

# --- Shift Breaks ---

@router.put("/shifts/{shift_id}/breaks", response_model=List[ShiftBreakRead])
async def bulk_update_breaks(
    shift_id: uuid.UUID,
    breaks: List[ShiftBreakCreate],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Replace all break windows for a shift."""
    service = AttendanceService(db)
    shift = await service.get_shift(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(shift.entity_id, current_user, db)
    
    updated_breaks = await service.replace_shift_breaks(shift_id, breaks)
    await db.commit()
    return updated_breaks

@router.get("/shifts/{shift_id}/breaks", response_model=List[ShiftBreakRead])
async def list_breaks(
    shift_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    return await service.get_shift_breaks(shift_id)

@router.delete("/shifts/{shift_id}/breaks/{break_id}")
async def delete_break(
    shift_id: uuid.UUID,
    break_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    shift = await service.get_shift(shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(shift.entity_id, current_user, db)
    
    await service.delete_shift_break(break_id)
    await db.commit()
    return {"message": "Break deleted successfully"}

# --- Roster ---

@router.get("/roster", response_model=List[RosterReadEnriched])
async def list_roster(
    entity_id: uuid.UUID,
    start_date: date,
    end_date: date,
    employment_id: Optional[uuid.UUID] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(entity_id, current_user, db)
    
    service = AttendanceService(db)
    return await service.get_roster_enriched(entity_id, start_date, end_date, employment_id)

@router.post("/roster/generate", response_model=List[ShiftRosterRead])
async def auto_generate_roster(
    data: RosterAutoGenerate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Smart roster generation using Employment rest_day and working_days_per_week."""
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(data.entity_id, current_user, db)
    
    service = AttendanceService(db)
    rosters = await service.auto_generate_roster(data)
    await db.commit()
    return rosters

@router.post("/roster/clear")
async def clear_roster(
    data: RosterClear,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Bulk clear roster for given employees and date range."""
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(data.entity_id, current_user, db)
    
    service = AttendanceService(db)
    deleted_count = await service.clear_roster(data)
    await db.commit()
    return {"message": f"Successfully cleared {deleted_count} roster entries", "deleted_count": deleted_count}

@router.post("/roster/bulk", response_model=List[ShiftRosterRead])
async def bulk_assign_roster(
    bulk_data: RosterBulkUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.api.v1.dependencies import get_entity_access
    await get_entity_access(bulk_data.entity_id, current_user, db)
    
    service = AttendanceService(db)
    rosters = await service.assign_roster_bulk(bulk_data)
    await db.commit()
    return rosters

@router.patch("/roster/{roster_id}", response_model=ShiftRosterRead)
async def update_roster_cell(
    roster_id: uuid.UUID,
    cell_update: RosterCellUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    roster = await service.update_roster_cell(roster_id, cell_update.shift_id, cell_update.day_type)
    if not roster:
        raise HTTPException(status_code=404, detail="Roster entry not found")
    await db.commit()
    return roster

@router.delete("/roster/{roster_id}")
async def delete_roster_cell(
    roster_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    deleted = await service.delete_roster_cell(roster_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Roster entry not found")
    await db.commit()
    return {"message": "Roster entry deleted"}

# --- Public Holidays ---

@router.get("/public-holidays", response_model=List[PublicHolidayRead])
async def list_public_holidays(
    entity_id: uuid.UUID,
    year: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    return await service.get_public_holidays(entity_id, year)

@router.post("/public-holidays", response_model=PublicHolidayRead)
async def create_public_holiday(
    data: PublicHolidayCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    ph = await service.create_public_holiday(data)
    await db.commit()
    return ph

@router.post("/public-holidays/seed", response_model=List[PublicHolidayRead])
async def seed_public_holidays(
    data: PublicHolidaySeed,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Seed all gazetted Singapore public holidays for a given year."""
    service = AttendanceService(db)
    holidays = await service.seed_sg_holidays(data.entity_id, data.year)
    await db.commit()
    return holidays

@router.patch("/public-holidays/{ph_id}", response_model=PublicHolidayRead)
async def update_public_holiday(
    ph_id: uuid.UUID,
    data: PublicHolidayUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    ph = await service.update_public_holiday(ph_id, data)
    if not ph:
        raise HTTPException(status_code=404, detail="Public holiday not found")
    await db.commit()
    return ph

@router.delete("/public-holidays/{ph_id}")
async def delete_public_holiday(
    ph_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    service = AttendanceService(db)
    deleted = await service.delete_public_holiday(ph_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Public holiday not found")
    await db.commit()
    return {"message": "Public holiday deleted"}
